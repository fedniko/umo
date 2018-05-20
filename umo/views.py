from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Alignment, Protection, Font, Side

from umo.forms import AddTeacherForm
from umo.models import *


# Create your views here.
class TeacherCreateView(CreateView):
    model = Person, Teacher
    fields = '__all__'


class TeacherUpdate(UpdateView):
    template_name = 'teacher_edit.html'
    success_url = reverse_lazy('teachers:list_teachers')
    model = Teacher
    fields = [
            'FIO',
            'Position',
            'Zvanie',
            'cathedra'
    ]
    labels = {
        'FIO': 'ФИО',
        'Position': 'Должность',
        'Zvanie': 'Звание',
        'cathedra': 'Кафедра'
    }


class TeacherDelete(DeleteView):
    model = Teacher
    success_url = reverse_lazy('teacher')


def list_teachers(request):
    all = Teacher.objects.all()
    return render(request, 'teachers_list.html', {'teachers': all})


def create_teacher(request):
     if request.method == 'POST':
        form = AddTeacherForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('teachers:list_teachers'))
        return render(request, 'teacher_form.html', {'form': form})
     form = AddTeacherForm()
     return render(request, 'teacher_form.html', {'form': form})


class StudentListView(ListView):
    model = GroupList
    context_object_name = 'student_list'
    success_url = reverse_lazy('student_changelist')
    template_name = "students_list.html"


class StudentCreateView(CreateView):
    model = GroupList
    fields = ['group']
    success_url = reverse_lazy('student_changelist')
    template_name = "student_form.html"

    def get_context_data(self, **kwargs):
        context = super(StudentCreateView, self).get_context_data(**kwargs)
        context['title'] = "Добавление студента"
        return context

    def form_valid(self, form):
        student_ = Student.objects.create()
        student_.FIO = form.data.get('fio')
        student_.StudentID = form.data.get('studid')
        student_.save()
        grouplist_ = form.save(commit=False)
        grouplist_.student = student_
        grouplist_.active = True
        grouplist_.save()
        return super().form_valid(form)


def student_delete(request):
    if request.method == 'POST':
        student_ = Student.objects.get(id = request.POST['item_id'])
        grouplist_ = GroupList.objects.get(student__id = student_.id)
        grouplist_.active = False
        return HttpResponseRedirect(reverse_lazy('student_changelist'))


class StudentUpdateView(UpdateView):
    model = GroupList
    fields = ['group']
    success_url = reverse_lazy('student_changelist')
    template_name = "student_form.html"
    context_object_name = 'student_list'

    def get_context_data(self, **kwargs):
        context = super(StudentUpdateView, self).get_context_data(**kwargs)
        context['title'] = "Изменение студента"
        return context

    def form_valid(self, form):
        student_ = self.object.student
        student_.FIO = form.data.get('fio')
        student_.StudentID = form.data.get('studid')
        student_.save()
        grouplist_ = self.object
        grouplist_.student = student_
        grouplist_.active = True
        grouplist_.save()
        return super().form_valid(form)


def delete_teacher(request):
    if request.method == 'POST':
        teacher_ = Teacher.objects.get(pk=request.POST['teacher'])
        teacher_.delete()
        return HttpResponseRedirect(reverse('teachers:list_teachers'))


def get_mark(str, value):
    if (str == 'Зачет'):
        if (value >= 60):
            return 'зач'
        else:
            return 'нз'
    else:
        if (value >= 85):
            return 'отл'
        elif (value >= 65):
            return 'хор'
        elif (value >= 55):
            return 'удовл'
        else:
            return 'неуд'


def get_mark_vedomost(str, inPoints, examPoints):
    value = inPoints + examPoints
    if (str == 'Зачет'):
        if (value >= 60):
            return 'Зачтено'
        else:
            return 'Не зачтено'
    else:
        if (inPoints < 45):
            return 'Не допущен'
        if (value >= 85):
            return 'Отлично'
        elif (value >= 65):
            return 'Хорошо'
        elif (value >= 55):
            return 'Удовлетворительно'
        else:
            return 'Неудовлетворительно'


def get_markSymbol(str, value):
    if (str == 'Зачет'):
        return None
    else:
        if (value >= 95):
            return 'A'
        elif (value >= 85):
            return 'B'
        elif (value >= 75):
            return 'C'
        elif (value >= 65):
            return 'D'
        elif (value >= 55):
            return 'E'
        elif (value >= 25):
            return 'FX'
        else:
            return 'F'


class BRSPointsListView(ListView):
    model = GroupList
    context_object_name = 'students_list'
    success_url = reverse_lazy('disciplines:disciplines_list')
    template_name = "brs_students.html"

    def get_queryset(self):
        disc = Discipline.objects.get(id = self.kwargs['pk'])
        return GroupList.objects.filter(group__program = disc.program)

    def get_context_data(self, **kwargs):
        context = super(BRSPointsListView, self).get_context_data(**kwargs)
        checkpoint = CheckPoint.objects.all()
        if (checkpoint.count() < 5):
            if (checkpoint.count() > 0):
                i = 0
                for ch in checkpoint:
                    i = i + 1
                    if (i == 1):
                        ch.Name = "Первый срез"
                    elif (i == 2):
                        ch.Name = "Второй срез"
                    elif (i == 3):
                        ch.Name = "Рубежный срез"
                    else:
                        ch.Name = "Рубежный срез с премиальными баллами"
                    ch.save()
            for i in range(checkpoint.count(), 5):
                ch = CheckPoint()
                if (i == 1):
                    ch.Name = "Первый срез"
                elif (i == 2):
                    ch.Name = "Второй срез"
                elif (i == 3):
                    ch.Name = "Рубежный срез"
                elif (i == 4):
                    ch.Name = "Рубежный срез с премиальными баллами"
                else:
                    ch.Name = "Всего баллов"
                ch.save()
        context['checkpoint'] = checkpoint
        discipline = Discipline.objects.get(id=self.kwargs['pk'])
        context['control_type'] = 'Баллы ' + discipline.control.controltype.lower()
        context['discipline'] = discipline
        context['grouplist'] = GroupList.objects.all()
        student = Student.objects.all()
        dict = {}
        for st in student:
            dict[str(st.id)] = {}
            dict[str(st.id)]['key'] = st.id
            i = 0
            for ch in checkpoint:
                i = i + 1
                newBRSpoints = BRSpoints.objects.filter(brs__discipline__id = discipline.id).filter(CheckPoint = ch).filter(student = st).first()
                if (newBRSpoints is None):
                    newBRSpoints = BRSpoints()
                    newBRSpoints.student = st
                    newBRSpoints.CheckPoint = ch
                    newBRSpoints.points = 0.0
                    newBRS = BRS.objects.filter(discipline__id = discipline.id).first()
                    if (newBRS is None):
                        newBRS = BRS()
                        newBRS.discipline = discipline
                        newBRS.eduperiod = EduPeriod.objects.all().first()
                        newBRS.semestr = DisciplineDetails.objects.filter(subject=discipline).first().semestr
                        newBRS.save()
                    newBRSpoints.brs = newBRS
                    newBRSpoints.save()
                dict[str(st.id)][str(i)] = newBRSpoints

            newExamMarks = ExamMarks.objects.filter(exam__discipline__id = discipline.id).filter(student = st).first()
            if (newExamMarks is None):
                newMarkSymbol = MarkSymbol.objects.filter(name='F').first()
                if (newMarkSymbol is None):
                    newMarkSymbol = MarkSymbol.objects.create(name='F')

                newMark = Mark.objects.filter(name='неуд').first()
                if (newMark is None):
                    newMark = Mark.objects.create(name='неуд')

                newExam = Exam.objects.filter(discipline__id = discipline.id).first()
                if (newExam is None):
                    newExam = Exam()
                    newControlType = ControlType.objects.filter(name=discipline.control.controltype).first()
                    if (newControlType is None):
                        newControlType = ControlType.objects.create(name = discipline.control.controltype)
                    newExam.controlType = newControlType
                    newExam.discipline = discipline
                    newExam.eduperiod = EduPeriod.objects.all().first()
                    newExam.examDate = 'не проставлена'
                    newExam.semestr = DisciplineDetails.objects.filter(subject=discipline).first().semestr
                    newExam.save()

                newExamMarks = ExamMarks()
                newExamMarks.student = st
                newExamMarks.inPoints = 0.0
                newExamMarks.examPoints = 0.0
                newExamMarks.markSymbol = newMarkSymbol
                newExamMarks.mark = newMark
                newExamMarks.exam = newExam
                newExamMarks.save()
            dict[str(st.id)]['6'] = newExamMarks
        context['dict'] = dict
        return context

    def post(self, request, *args, **kwargs):
        if (request.POST.get('save')):
            studid = request.POST.getlist('studid')
            points = []
            points.append(request.POST.getlist('points1'))
            points.append(request.POST.getlist('points2'))
            points.append(request.POST.getlist('points3'))
            points.append(request.POST.getlist('points4'))
            points.append(request.POST.getlist('points6'))
            arr_size = len(studid)
            checkpoint = CheckPoint.objects.all()
            discipline = Discipline.objects.get(id=self.kwargs['pk'])
            for i in range(0, arr_size):
                st = Student.objects.get(id=studid[i])
                k = 0

                exammarks = ExamMarks.objects.filter(exam__discipline__id=discipline.id).get(student=st)
                exammarks.examPoints = float(points[4][i].replace(',', '.'))
                exammarks.inPoints = float(points[3][i].replace(',', '.'))

                totalPoints = exammarks.examPoints + exammarks.inPoints

                for ch in checkpoint:
                    brspoints = BRSpoints.objects.filter(brs__discipline__id=discipline.id).filter(CheckPoint=ch).get(
                        student=st)
                    if (k != 4):
                        brspoints.points = float(points[k][i].replace(',', '.'))
                        k = k + 1
                    else:
                        brspoints.points = totalPoints
                        k = 0
                    brspoints.save()

                tempMarkSymbol = get_markSymbol(discipline.control.controltype, totalPoints)
                tempMark = get_mark(discipline.control.controltype, totalPoints)

                if (tempMarkSymbol is None):
                    newMarkSymbol = None
                else:
                    newMarkSymbol = MarkSymbol.objects.filter(name=tempMarkSymbol).first()
                    if (newMarkSymbol is None):
                        newMarkSymbol = MarkSymbol.objects.create(name=tempMarkSymbol)

                newMark = Mark.objects.filter(name=tempMark).first()
                if (newMark is None):
                    newMark = Mark.objects.create(name=tempMark)

                exammarks.markSymbol = newMarkSymbol
                exammarks.mark = newMark
                exammarks.save()

            return HttpResponseRedirect(reverse('brs_studentlist', args=(self.kwargs['pk'])))

        elif request.POST.get('vedomost'):
            # определяем стили
            font_main = Font(name='Times New Roman',
                             size=12,
                             bold=False,
                             italic=False,
                             vertAlign=None,
                             underline='none',
                             strike=False,
                             color='FF000000')

            font_bold = Font(name='Times New Roman',
                             size=12,
                             bold=True,
                             italic=False,
                             vertAlign=None,
                             underline='none',
                             strike=False,
                             color='FF000000')

            font_bold_s = Font(name='Times New Roman',
                               size=10,
                               bold=True,
                               italic=False,
                               vertAlign=None,
                               underline='none',
                               strike=False,
                               color='FF000000')

            font_calibri = Font(name='Calibri',
                                size=11,
                                bold=False,
                                italic=False,
                                vertAlign=None,
                                underline='none',
                                strike=False,
                                color='FF000000')

            font_arial = Font(name='Arial Cyr',
                              size=12,
                              bold=False,
                              italic=True,
                              vertAlign=None,
                              underline='none',
                              strike=False,
                              color='FF000000')

            fill = PatternFill(fill_type='solid',
                               start_color='c1c1c1',
                               end_color='c2c2c2')

            border = Border(left=Side(border_style='thin',
                                      color='FF000000'),
                            right=Side(border_style='thin',
                                       color='FF000000'),
                            top=Side(border_style='thin',
                                     color='FF000000'),
                            bottom=Side(border_style='thin',
                                        color='FF000000'),
                            diagonal=Side(border_style='thin',
                                          color='FF000000'),
                            diagonal_direction=0,
                            outline=Side(border_style='thin',
                                         color='FF000000'),
                            vertical=Side(border_style='thin',
                                          color='FF000000'),
                            horizontal=Side(border_style='thin',
                                            color='FF000000')
                            )
            align_center = Alignment(horizontal='center',
                                     vertical='center',
                                     text_rotation=0,
                                     wrap_text=False,
                                     shrink_to_fit=False,
                                     indent=0)
            align_center2 = Alignment(horizontal='center',
                                      vertical='center',
                                      text_rotation=0,
                                      wrap_text=True,
                                      shrink_to_fit=False,
                                      indent=0)
            align_left = Alignment(horizontal='left',
                                   vertical='center',
                                   text_rotation=0,
                                   wrap_text=False,
                                   shrink_to_fit=False,
                                   indent=0)
            number_format = 'General'
            protection = Protection(locked=True,
                                    hidden=False)

            # объект
            wb = Workbook()

            # активный лист
            ws = wb.active

            # название страницы
            # ws = wb.create_sheet('первая страница', 0)
            ws.title = 'первая страница'

            # объединение ячеек
            ws.merge_cells('A1:I1')
            ws.merge_cells('A2:I2')
            ws.merge_cells('A3:I3')
            ws.merge_cells('A5:B5')
            ws.merge_cells('A6:B6')
            ws.merge_cells('C6:D6')
            ws.merge_cells('A7:B7')
            ws.merge_cells('C7:G7')
            ws.merge_cells('A8:C8')
            ws.merge_cells('D8:G8')
            ws.merge_cells('A9:B9')
            ws.merge_cells('C9:D9')
            ws.merge_cells('E39:F39')
            ws.merge_cells('E40:F40')
            ws.merge_cells('E41:F41')
            ws.merge_cells('E42:F42')
            ws.merge_cells('E43:F43')
            ws.merge_cells('E44:F44')
            ws.merge_cells('E45:F45')
            ws.merge_cells('E46:F46')
            ws.merge_cells('G39:H39')
            ws.merge_cells('G40:H40')
            ws.merge_cells('G41:H41')
            ws.merge_cells('G42:H42')
            ws.merge_cells('G43:H43')
            ws.merge_cells('G44:H44')
            ws.merge_cells('G45:H45')
            ws.merge_cells('G46:H46')
            ws.merge_cells('B49:C49')
            ws.merge_cells('D49:E49')

            # данные для строк
            group_name = str(request.POST.get('selected_group'))
            disc_id = self.kwargs['pk']
            exam = Exam.objects.get(discipline__id=disc_id)
            studid = request.POST.getlist('studid')

            inpoints = request.POST.getlist('points4')
            exampoints = request.POST.getlist('points6')
            arr_size = len(studid)

            _row = 12
            _column = 4
            k = 1

            zachteno = 0
            ne_zachteno = 0
            ne_attest = 0
            otl = 0
            horosho = 0
            udovl = 0
            neudovl = 0
            ne_yavka = 0

            ws.cell(row=1, column=1).value = 'ФГАОУ ВО «Северо-Восточный федеральный университет им.М.К.Аммосова'
            ws.cell(row=2, column=1).value = 'Институт математики и информатики'
            ws.cell(row=3, column=1).value = 'Ведомость текущей и промежуточной аттестации'
            ws.cell(row=5, column=1).value = 'Семестр: ' + str(
                exam.semestr.name) + ', ' + exam.eduperiod.beginyear + '-' + exam.eduperiod.endyear + ' уч.г.'
            ws.cell(row=6, column=1).value = 'Форма контроля:'
            ws.cell(row=6, column=3).value = exam.controlType.name
            ws.cell(row=6, column=5).value = 'курс 1'
            ws.cell(row=6, column=6).value = 'группа:'
            ws.cell(row=6, column=7).value = group_name
            ws.cell(row=7, column=1).value = 'Дисциплина:'
            ws.cell(row=7, column=3).value = exam.discipline.Name
            ws.cell(row=8, column=1).value = 'Фамилия, имя, отчество преподавателя:'
            ws.cell(row=8, column=4).value = exam.discipline.lecturer.FIO
            ws.cell(row=9, column=1).value = 'Дата проведения зачета/экзамена:'
            ws.cell(row=9, column=3).value = exam.examDate
            ws.cell(row=11, column=1).value = '№'
            ws.cell(row=11, column=2).value = 'Фамилия, имя, отчество'
            ws.cell(row=11, column=3).value = '№ зачетной книжки'
            ws.cell(row=11, column=4).value = 'Сумма баллов за текущую работу-рубеж.срез'
            ws.cell(row=11, column=5).value = 'Баллы ' + exam.controlType.name + ' (бонусные баллы)'
            ws.cell(row=11, column=6).value = 'Всего баллов'
            ws.cell(row=11, column=7).value = 'Оценка прописью'
            ws.cell(row=11, column=8).value = 'Буквенный эквивалент'
            ws.cell(row=11, column=9).value = 'Подпись преподавателя'
            ws.cell(row=39, column=2).value = 'зачтено'
            ws.cell(row=40, column=2).value = 'не зачтено'
            ws.cell(row=41, column=2).value = 'не аттест'
            ws.cell(row=42, column=2).value = '5(отлично)'
            ws.cell(row=43, column=2).value = '4(хорошо)'
            ws.cell(row=44, column=2).value = '3(удовл)'
            ws.cell(row=45, column=2).value = '2(неудовл)'
            ws.cell(row=46, column=2).value = 'не явка'
            ws.cell(row=39, column=5).value = 'Сумма баллов'
            ws.cell(row=40, column=5).value = '95-100'
            ws.cell(row=41, column=5).value = '85-94,9'
            ws.cell(row=42, column=5).value = '75-84,9'
            ws.cell(row=43, column=5).value = '65-74,9'
            ws.cell(row=44, column=5).value = '55-64,9'
            ws.cell(row=45, column=5).value = '25-54,9'
            ws.cell(row=46, column=5).value = '0-24,9'
            ws.cell(row=39, column=7).value = 'Буквенный эквивалент оценки'
            ws.cell(row=40, column=7).value = 'A'
            ws.cell(row=41, column=7).value = 'B'
            ws.cell(row=42, column=7).value = 'C'
            ws.cell(row=43, column=7).value = 'D'
            ws.cell(row=44, column=7).value = 'E'
            ws.cell(row=45, column=7).value = 'FX'
            ws.cell(row=46, column=7).value = 'F'
            ws.cell(row=49, column=2).value = 'Директор ИМИ СВФУ____________________'
            ws.cell(row=49, column=4).value = 'В.И.Афанасьева'

            for i in range(0, arr_size):
                gl = Student.objects.get(id=studid[i])
                ws.cell(row=_row, column=1).value = str(k)
                k += 1
                ws.cell(row=_row, column=2).value = gl.FIO
                ws.cell(row=_row, column=3).value = gl.StudentID
                ws.cell(row=_row, column=_column).value = str(float(inpoints[i].replace(',', '.'))).replace('.', ',')
                ws.cell(row=_row, column=_column + 1).value = str(float(exampoints[i].replace(',', '.'))).replace('.', ',')
                totalpoints = float(inpoints[i].replace(',', '.')) + float(exampoints[i].replace(',', '.'))
                ws.cell(row=_row, column=_column + 2).value = str(totalpoints).replace('.', ',')
                ws.cell(row=_row, column=_column + 3).value = get_mark_vedomost(exam.controlType.name, float(inpoints[i].replace(',', '.')), float(exampoints[i].replace(',', '.')))
                ws.cell(row=_row, column=_column + 4).value = get_markSymbol(exam.controlType.name, totalpoints)
                _row += 1

            for cellObj in ws['G12:G37']:
                for cell in cellObj:
                    if ws[cell.coordinate].value == 'Зачтено':
                        zachteno = zachteno + 1
                    elif ws[cell.coordinate].value == 'Не зачтено':
                        ne_zachteno = ne_zachteno + 1
                    elif ws[cell.coordinate].value == 'Не аттест':
                        ne_attest = ne_attest + 1
                    elif ws[cell.coordinate].value == 'Отлично':
                        otl = otl + 1
                    elif ws[cell.coordinate].value == 'Хорошо':
                        horosho = horosho + 1
                    elif ws[cell.coordinate].value == 'Удовлетворительно':
                        udovl = udovl + 1
                    elif ws[cell.coordinate].value == 'Неудовлетворительно':
                        neudovl = neudovl + 1
                    elif ws[cell.coordinate].value == 'Не явка':
                        ne_yavka = ne_yavka + 1

            ws.cell(row=39, column=3).value = str(zachteno)
            ws.cell(row=40, column=3).value = str(ne_zachteno)
            ws.cell(row=41, column=3).value = str(ne_attest)
            ws.cell(row=42, column=3).value = str(otl)
            ws.cell(row=43, column=3).value = str(horosho)
            ws.cell(row=44, column=3).value = str(udovl)
            ws.cell(row=45, column=3).value = str(neudovl)
            ws.cell(row=46, column=3).value = str(ne_yavka)

            # шрифты
            for cellObj in ws['A1:I37']:
                for cell in cellObj:
                    ws[cell.coordinate].font = font_main

            for cellObj in ws['G12:G37']:
                for cell in cellObj:
                    ws[cell.coordinate].font = font_bold_s

            for cellObj in ws['B12:B37']:
                for cell in cellObj:
                    ws[cell.coordinate].font = font_calibri

            for cellObj in ws['H12:H37']:
                for cell in cellObj:
                    ws[cell.coordinate].font = font_calibri

            for cellObj in ws['E12:E37']:
                for cell in cellObj:
                    ws[cell.coordinate].font = font_bold

            for cellObj in ws['E11:I11']:
                for cell in cellObj:
                    ws[cell.coordinate].font = Font(name='Times New Roman',
                                                  size=9,
                                                  bold=False,
                                                  italic=False,
                                                  vertAlign=None,
                                                  underline='none',
                                                  strike=False,
                                                  color='FF000000')

            ws['A3'].font = font_bold
            ws['C7'].font = font_bold
            ws['D8'].font = font_bold
            ws['F6'].font = font_bold
            ws['C7'].font = font_arial
            ws['D8'].font = font_arial
            ws['G6'].font = Font(name='Arial Cyr',
                                 size=12,
                                 bold=False,
                                 italic=True,
                                 vertAlign=None,
                                 underline='single',
                                 strike=False,
                                 color='FF000000')
            ws['C9'].font = Font(name='Calibri',
                                 size=11,
                                 bold=False,
                                 italic=False,
                                 vertAlign=None,
                                 underline='single',
                                 strike=False,
                                 color='FF000000')
            ws['A11'].font = Font(name='Times New Roman',
                                  size=10,
                                  bold=False,
                                  italic=False,
                                  vertAlign=None,
                                  underline='none',
                                  strike=False,
                                  color='FF000000')
            ws['B11'].font = Font(name='Times New Roman',
                                  size=10,
                                  bold=False,
                                  italic=False,
                                  vertAlign=None,
                                  underline='none',
                                  strike=False,
                                  color='FF000000')
            ws['C11'].font = Font(name='Times New Roman',
                                  size=9,
                                  bold=False,
                                  italic=False,
                                  vertAlign=None,
                                  underline='none',
                                  strike=False,
                                  color='FF000000')
            ws['D11'].font = Font(name='Times New Roman',
                                  size=8,
                                  bold=False,
                                  italic=False,
                                  vertAlign=None,
                                  underline='none',
                                  strike=False,
                                  color='FF000000')
            ws['C6'].font = Font(name='Times New Roman',
                                 size=14,
                                 bold=False,
                                 italic=True,
                                 vertAlign=None,
                                 underline='single',
                                 strike=False,
                                 color='FF000000')

            # увеличиваем все строки по высоте
            max_row = ws.max_row
            i = 1
            while i <= max_row:
                rd = ws.row_dimensions[i]
                rd.height = 16
                i += 1

            # вручную устанавливаем высоту первой строки
            rd = ws.row_dimensions[11]
            rd.height = 48

            # сетка
            for cellObj in ws['A11:I37']:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    ws[cell.coordinate].border = border

            for cellObj in ws['B39:C46']:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    ws[cell.coordinate].border = border

            for cellObj in ws['E39:H46']:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    ws[cell.coordinate].border = border

            # выравнивание
            for cellObj in ws['A1:I37']:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    ws[cell.coordinate].alignment = align_center

            for cellObj in ws['A11:I11']:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    ws[cell.coordinate].alignment = align_center2

            for cellObj in ws['A5:I9']:
                for cell in cellObj:
                    # print(cell.coordinate, cell.value)
                    ws[cell.coordinate].alignment = align_left

            # перетягивание ячеек
            dims = {}
            for cellObj in ws['G11:G37']:
                for cell in cellObj:
                    if cell.value:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
            for col, value in dims.items():
                # value * коэфициент
                ws.column_dimensions[col].width = value * 1.5

            dims = {}
            for cellObj in ws['A11:A37']:
                for cell in cellObj:
                    if cell.value:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
            for col, value in dims.items():
                # value * коэфициент
                ws.column_dimensions[col].width = value * 3

            dims = {}
            for cellObj in ws['B11:B37']:
                for cell in cellObj:
                    if cell.value:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
            for col, value in dims.items():
                # value * коэфициент
                ws.column_dimensions[col].width = value * 1.5

            dims = {}
            for cellObj in ws['D11:D37']:
                for cell in cellObj:
                    if cell.value:
                        dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
            for col, value in dims.items():
                # value * коэфициент
                ws.column_dimensions[col].width = value * 0.25

            # сохранение файла в выбранную директорию
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=vedomost.xlsx'

            wb.save(response)

            return response